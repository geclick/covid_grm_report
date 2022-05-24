from openpyxl import load_workbook
from zipfile import BadZipFile

from django import forms
from django.core.validators import FileExtensionValidator

from . import models, validators

HIGIENE_EXCEL_COLUMMS = [
    "CÓDIGO",
    "ID PROVINCIA",
    "NOMBRE Y APELLIDOS",
    "EDAD",
    "SEXO",
    "C IDENTIDAD/ PASAPORTE",
    "DIRECCION",
    "AREA DE SALUD",
    "MUNICIPIO",
    "PROVINCIA",
    "CONDICIÓN",
    "PAÍS DE PROCEDENCIA",
    "F INICIO SÍNTOMAS",
    "FECHA DE TOMA DE MUESTRA",
    "TIPO DE MUESTRA",
    "PROCEDENCIA DE LA ENTREGA",
    "RESULTADO",
    "FECHA/ INICIO SÍNTOMAS",
    "CT",
    "FECHA DE SALIDA",
]

STG_EXCEL_COLUMMS = [
    "CODIGO LBM",
    "ID MUESTRA DEL AREA",
    "FECHA",
    "NOMBRE Y APELLIDOS",
    "EDAD",
    "SEXO",
    "C IDENTIDAD/ PASAPORTE",
    "DIRECCIÓN",
    "AREA DE SALUD",
    "MUNICIPIO",
    "PROVINCIA",
    "CONDICIÓN",
    "PAIS DE PROCEDENFCIA",
    "F INICIO SÍNTOMAS",
    "FECHA DE TOMA DE MUESTRA",
    "TIPO DE MUESTRA",
    "PROCEDENFCIA DE LA ENFTREGA",
    "RESULTADO",
    "CUÁL",
    "CT",
    "Numero de Placa",
    "FECHA DE RESULTADO",
]

GTM_EXCEL_COLUMMS = [
    "CÓD LBM",
    "CÓD AREA SALUD",
    "FECHA RECEP. LBM",
    "NOMBRE Y APELLIDOS",
    "EDAD",
    "D/M/A",
    "SEXO",
    "COLOR DE PIEL",
    "C IDENTIDAD/ PASAP0RTE",
    "DIRECCIÓN",
    "AREA DE SALUD",
    "MUNICIPIO",
    "PROVINCIA",
    "CONDICIÓN",
    "PAIS DE PROCEDENCIA",
    "F. I. S.",
    "FECHA T.M",
    "FIEBRE",
    "RINORREA",
    "CONGESTION NASAL",
    "TOS",
    "EXPECTORACION",
    "DOLOR DE GARGANTA",
    "CEFALEA",
    "DIFICULTAD RESPIRATORIA",
    "OTROS (ESPECIFICAR)",
    "TIPO DE MUESTRA",
    "PROCEDENCIA DE LA ENTREGA",
    "RESULTADO",
    "CT",
    "OBSERVACIÓN",
    "FECHA RESUL",
    "PLACA NÚMERO",
    "CONTACTO DEL CASO CONFIRMADO",
    "RELACION CON EVENTO",
    "TOTAL DE CONTACTOS",
]

HOL_EXCEL_COLUMMS = [
    "R",
    "Priorizado",
    "NUMERO EN EL LAB MHLG",
    "ID DE LA PROVINCIA",
    "FECHA ENTRADA DE LAS MUESTRAS",
    "ESTADO DE LAS MUESTRAS",
    "HORA DE ENTRADA",
    "INDICADO POR",
    "NOMBRES Y APELLIDOS",
    "Edad",
    "Sexo",
    "d/M/a",
    "Direccion",
    "A/Salud",
    "MUNICIPIO",
    "PROVINCIA",
    "MOTIVO",
    "PAIS DE PROCEDENCIA",
    "FIS",
    "FTM",
    "Tpo Muestra",
    "Prov./procedencia de la Muestra",
    "PCR",
    "CT",
    "Fecha Resultado",
    "Observaciones",
]


GRM_EXCEL_COLUMMS = [
    "R",
    "PRIORIDAD",
    "NUMERO EN EL LAB MGRM",
    "ID DE LA PROVINCIA",
    "FECHA ENTRADA DE LAS MUESTRAS",
    "ESTADO DE LAS MUESTRAS",
    "HORA DE ENTRADA",
    "INDICADO POR",
    "NOMBRES Y APELLIDOS",
    "Edad",
    "Sexo",
    "d/M/a",
    "Direccion",
    "A/Salud",
    "MUNICIPIO",
    "PROVINCIA",
    "MOTIVO",
    "PAIS DE PROCEDENCIA",
    "FIS",
    "FTM",
    "Tpo Muestra",
    "Prov./procedencia de la Muestra",
    "PCR",
    "CT",
    "Fecha Resultado",
    "NO. DE PLACA",
    "Observaciones",
]

LTU_EXCEL_COLUMMS = [
    "No",
    "Número",
    "Nombres y Apellidos",
    "Edad",
    "Sexo",
    "Dirección",
    "Grupo pesquisa",
    "Síntomas si",
    "Síntomas no",
    "FIS",
    "FTM",
    "A Salud",
    "Municipio",
    "Hospital",
    "Procedencia",
    "Resultados",
]

CMG_EXCEL_COLUMMS = [
    "NO. DE CORRIDAS",
    "CÓDIGO",
    "ID PROV.",
    "NOMBRE Y APELLIDOS",
    "EDAD",
    "SEXO",
    "C IDENTIDAD/ PASAPORTE",
    "DIRECCIÓN",
    "AREA DE SALUD",
    "MUNICIPIO",
    "PROVINCIA",
    "CONDICIÓN",
    "PAIS DE PROCEDENCIA",
    "FIS",
    "FECHA DE TOMA DE MUESTRA",
    "TIPO DE MUESTRA",
    "PROCEDENCIA DE LA MUESTRA",
    "RESULTADO",
    "CT",
    "FECHA DE SALIDA",
]

CAV_EXCEL_COLUMMS = [
    "CÓDIGO",
    "ID PROVINCIA",
    "NOMBRE Y APELLIDOS",
    "EDAD",
    "SEXO",
    "C IDENTIDAD/ PASAPORTE",
    "DIRECCIÓN",
    "AREA DE SALUD",
    "MUNICIPIO",
    "PROVINCIA",
    "CONDICIÓN",
    "PAIS DE PROCEDENCIA",
    "F INICIO SÍNTOMAS",
    "FECHA DE TOMA DE MUESTRA",
    "TIPO DE MUESTRA",
    "PROCEDENCIA DE LA ENTREGA",
    "RESULTADO",
    "CT",
    "FECHA DE SALIDA",
    "OBSERVACIONES",
    "SALA",
]

CFG_EXCEL_COLUMMS = [
    "CORRIDA",
    "CODIGO",
    "ID PROVINCIA",
    "NOMBRE Y APELLIDOS",
    "EDAD",
    "SEXO",
    "CARNET",
    "DIRECCIÓN",
    "AREA DE SALUD",
    "MUNICIPIO",
    "PROVINCIA",
    "CONDICIÓN",
    "PAIS DE PROCEDENCIA",
    "F INICIO SÍNTOMAS",
    "FECHA DE TOMA DE MUESTRA",
    "TIPO DE MUESTRA",
    "PROCEDENCIA DE LA ENTREGA",
    "RESULTADO",
    "CT GEN ORF",
    "ct gen E",
    "CT GEN N",
    "FECHA DE SALIDA",
    "FECHAENTRADA",
]

VCL_EXCEL_COLUMMS = [
    "ID PROVINCIA",
    "NOMBRE Y APELLIDOS",
    "EDAD",
    "SEXO",
    "C IDENTIDAD/ PASAPORTE",
    "DIRECCIÓN",
    "AREA DE SALUD",
    "MUNICIPIO",
    "PROVINCIA",
    "CONDICIÓN",
    "PAIS DE PROCEDENCIA",
    "F INICIO SÍNTOMAS",
    "FECHA DE TOMA DE MUESTRA",
    "TIPO DE MUESTRA",
    "PROCEDENCIA DE LA ENTREGA",
    "RESULTADOS",
    "CT",
    "FECHA DE SALIDA",
]

MTZ_EXCEL_COLUMMS = [
    "No. Muestra",
    "No. Paquete",
    "VIP",
    "Tipo de sujeto",
    "Nombre(s) y Apellidos",
    "No. Carné de Identidad",
    "No. Pasaporte",
    "Edad",
    "Sexo",
    "Provincia",
    "Municipio",
    "Teléfono de contacto",
    "E-mail de contacto",
    "Dirección particular",
    "Area de salud",
    "Hospital",
    "Centro de aislamiento",
    "Laboratorio",
    "Punto de entrada al país",
    "País de procedencia",
    "Tipo de estudio",
    "Fecha de la toma de la muestra",
    "Tipo de muestra",
    "Impresión diagnóstica",
    "Fecha de inicio de los síntomas",
    "RESULTADO",
]

IPK_EXCEL_COLUMMS = [
    "Id IPK",
    "Id Prov",
    "Nombres y Apellidos",
    "Tipo de Estudio",
    "CI o PASAPORTE",
    "Edad",
    "Provincia",
    "Municipios",
    "Hospitales",
    "Dirección Particular",
    "Fecha entrada Muestra al Laborat",
    "Resultado_f",
    "Valor CT_f",
    "Gen_f",
    "Parte_f",
    "Fecha Salida_F",
]

HAB_EXCEL_COLUMMS = [
    "#/CANT",
    "CÓDIGO HODFR",
    "ID PROVINCIA",
    "NOMBRE Y APELLIDOS",
    "EDAD",
    "SEXO",
    "C IDENTIDAD/ PASAPORTE",
    "DIRECCIÓN",
    "AREA DE SALUD",
    "MUNICIPIO",
    "PROVINCIA",
    "CONDICIÓN",
    "PAIS DE PROCEDENCIA",
    "F INICIO SÍNTOMAS",
    "FECHA DE TOMA DE MUESTRA",
    "TIPO DE MUESTRA",
    "PROCEDENCIA DE LA ENTREGA",
    "RESULTADO",
    "CT",
    "FECHA DE SALIDA",
]

ART_EXCEL_COLUMMS = [
    "No. Muestra",
    "No. Paquete",
    "VIP",
    "Tipo de sujeto",
    "Nombre(s) y Apellidos",
    "No. Carné de Identidad",
    "No. Pasaporte",
    "Edad",
    "Sexo",
    "Provincia",
    "Municipio",
    "Teléfono de contacto",
    "E-mail de contacto",
    "Dirección particular",
    "Area de salud",
    "Hospital",
    "Centro de aislamiento",
    "Laboratorio",
    "Punto de entrada al país",
    "País de procedencia",
    "Tipo de estudio",
    "Fecha de la toma de la muestra",
    "Tipo de muestra",
    "Impresión diagnóstica",
    "Fecha de inicio de los síntomas",
    "Resultado",
    "CT",
]

PR_EXCEL_COLUMMS = [
    "NUM LBMPR",
    "NUM DIA",
    "FECHA DIA",
    "COD AREA",
    "NOMBRE Y APELLIDOS",
    "EDAD",
    "SEXO",
    "C IDENTIDAD PASAPORTE",
    "DIRECCIÓN",
    "AREA DE SALUD",
    "MUNICIPIO",
    "PROVINCIA",
    "CONDICIÓN",
    "PAIS DE PROCEDENCIA",
    "F INICIO SÍNTOMAS",
    "FECHA DE TOMA DE MUESTRA",
    "TIPO DE MUESTRA",
    "PROCEDENCIA DE LA ENTREGA",
    "FECHA ENTRADA AL LABORATORIO",
    "RESULTADO",
    "CT:N",
    "CT:ORF",
    "CT:E",
    "FECHA DE SALIDA",
    "PLACA",
]

LAB_MATCHED_LIST = {
    "HABANA": HAB_EXCEL_COLUMMS,
    "SANTIAGO": STG_EXCEL_COLUMMS,
    "ARTEMISA": ART_EXCEL_COLUMMS,
    "CAMAGUEY": CMG_EXCEL_COLUMMS,
    "CIEGODEAVILA": CAV_EXCEL_COLUMMS,
    "CIENFUEGOS": CFG_EXCEL_COLUMMS,
    "GUANTANAMO": GTM_EXCEL_COLUMMS,
    "HOLGUIN": HOL_EXCEL_COLUMMS,
    "IPK": IPK_EXCEL_COLUMMS,
    "MATANZAS": MTZ_EXCEL_COLUMMS,
    "PINARDELRIO": PR_EXCEL_COLUMMS,
    "TUNAS": LTU_EXCEL_COLUMMS,
    "VILLACLARA": VCL_EXCEL_COLUMMS,
    "GRANMA": GRM_EXCEL_COLUMMS,
}


class EpidemiologiaUploadForm(forms.ModelForm):
    archivo = forms.FileField(
        validators=[FileExtensionValidator(allowed_extensions=["xlsx"])]
    )

    class Meta:
        model = models.EpidemiologiaUploadModel
        fields = "__all__"


class HigieneUploadForm(forms.ModelForm):
    archivo = forms.FileField(
        validators=[
            FileExtensionValidator(allowed_extensions=["xlsx"]),
            validators.validate_file_size,
        ]
    )

    class Meta:
        model = models.HigieneUploadModel
        fields = "__all__"

    def clean(self):
        file = self.cleaned_data.get("archivo", None)

        if file:

            try:
                excel = load_workbook(file)
                sheet = excel.active

                min_column = sheet.min_column
                min_row = sheet.min_row
                max_column = min_row + (len(HIGIENE_EXCEL_COLUMMS) - 1)

                for row in sheet.iter_rows(
                    min_row=min_row,
                    max_row=min_row,
                    min_col=min_column,
                    max_col=max_column,
                ):
                    for cell in row:
                        if cell.value and (
                            cell.value.strip() not in HIGIENE_EXCEL_COLUMMS
                        ):
                            # print(cell.value)
                            self.add_error(
                                "archivo",
                                forms.ValidationError(
                                    "El encabezado {} de la celda {} no se corresponde con ningún valor esperado".format(  # noqa: E501
                                        cell.value, cell.coordinate
                                    )
                                ),
                            )

            except BadZipFile:
                self.add_error(
                    "archivo",
                    forms.ValidationError("Este archivo no es un excel."),
                )

        return self.cleaned_data


class ResultModelForm(forms.ModelForm):
    archivo = forms.FileField(
        validators=[
            FileExtensionValidator(allowed_extensions=["xlsx"]),
            validators.validate_file_size,
        ]
    )

    class Meta:
        model = models.ResultModel
        fields = "__all__"

    def clean(self):
        file = self.cleaned_data.get("archivo", None)
        laboratorio = self.cleaned_data.get("lab_procesado", None)

        if file and laboratorio:

            try:
                excel = load_workbook(file)
                sheet = excel.active

                min_column = sheet.min_column
                min_row = sheet.min_row

                try:

                    column_list = LAB_MATCHED_LIST[laboratorio.nombre.upper()]

                    max_column = min_row + (len(column_list) - 1)

                    for row in sheet.iter_rows(
                        min_row=min_row,
                        max_row=min_row,
                        min_col=min_column,
                        max_col=max_column,
                    ):
                        for cell in row:
                            if cell.value and (
                                cell.value.strip() not in column_list
                            ):
                                # print(cell.value)
                                self.add_error(
                                    "archivo",
                                    forms.ValidationError(
                                        "El encabezado {} de la celda {} no se corresponde con ningún valor esperado para el laboratorio {}".format(  # noqa: E501
                                            cell.value,
                                            cell.coordinate,
                                            laboratorio.nombre,
                                        )
                                    ),
                                )
                except (KeyError, AttributeError):
                    pass

            except BadZipFile:
                self.add_error(
                    "archivo",
                    forms.ValidationError("Este archivo no es un excel."),
                )

        return self.cleaned_data
